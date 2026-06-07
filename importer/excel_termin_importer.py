import csv
from pathlib import Path

from openpyxl import load_workbook


class ExcelTerminImporter:
    def load_rows(self, file_path: str, sheet_name: str | None = None) -> list[list[str]]:
        path = Path(file_path)

        if path.suffix.lower() == ".csv":
            return self._load_csv_rows(path)

        if path.suffix.lower() == ".xlsx":
            return self._load_xlsx_rows(path, sheet_name=sheet_name)

        raise ValueError(f"Nicht unterstütztes Dateiformat: {path.suffix}")

    def _load_csv_rows(self, file_path: Path) -> list[list[str]]:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, delimiter=";")
            return [[self._normalize_cell(cell) for cell in row] for row in reader]

    def _load_xlsx_rows(self, file_path: Path, sheet_name: str | None = None) -> list[list[str]]:
        workbook = load_workbook(filename=file_path, data_only=True)

        if sheet_name and sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows: list[list[str]] = []

        for row in sheet.iter_rows(values_only=True):
            rows.append([self._normalize_cell(cell) for cell in row])

        return rows

    def _normalize_cell(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()
